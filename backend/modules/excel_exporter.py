from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


def export_to_excel(df, output_path):

    # Create Excel
    df.to_excel(
        output_path,
        index=False,
        sheet_name="PMC Report"
    )

    workbook = load_workbook(output_path)

    worksheet = workbook["PMC Report"]

    # Header style
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_side = Side(
        style="thin",
        color="D9E1F2"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    # Header formatting
    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = border

    # Body formatting
    for row in worksheet.iter_rows():

        for cell in row:

            cell.border = border

            cell.alignment = Alignment(
                vertical="center"
            )

    # Column widths
    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            40
        )

    # Freeze first row
    worksheet.freeze_panes = "A2"

    # Enable filter
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    workbook.save(output_path)